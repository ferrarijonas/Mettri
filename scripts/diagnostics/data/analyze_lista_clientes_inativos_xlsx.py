#!/usr/bin/env python3
"""
Lista_ClientesInativos xlsx -> Retomar lastActivityByChat.
Colunas: Nome | Telefone | Ultimo Pedido | Ticket Medio.
chatId vem do app (match telefone + aliases BR). Ver spec retomar secao 10.
"""
import re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

def digits_only(s):
    return re.sub(r"\D+", "", s or "")

def main():
    default = Path.home() / "Desktop" / "Lista_ClientesInativos_24-03-2026-16-14.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not path.is_file():
        print("Ficheiro nao encontrado:", path)
        sys.exit(1)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    print("=== Cabecalho ===")
    for i, h in enumerate(header):
        print(f"  [{i}] {h!r}")
    idx_nome, idx_tel, idx_pedido = 0, 1, 2
    data_rows = rows[1:]
    by_phone = defaultdict(list)
    empty_phone = 0
    bad_date = 0
    for i, row in enumerate(data_rows, start=2):
        nome = str((row[idx_nome] if len(row) > idx_nome else "") or "").strip()
        tel_s = str((row[idx_tel] if len(row) > idx_tel else "") or "").strip()
        d = row[idx_pedido] if len(row) > idx_pedido else None
        dig = digits_only(tel_s)
        if not dig:
            empty_phone += 1
        dt = d if isinstance(d, datetime) else None
        if d is not None and not isinstance(d, datetime) and str(d).strip():
            bad_date += 1
        if dig:
            by_phone[dig].append((i, nome, dt, tel_s))
    dup = {k: v for k, v in by_phone.items() if len(v) > 1}
    print("\n=== Resumo ===")
    print("  Linhas dados:", len(data_rows))
    print("  Telefones unicos:", len(by_phone))
    print("  Sem digitos telefone:", empty_phone)
    print("  Grupos telefone duplicado:", len(dup))
    print("  Pedido nao-datetime:", bad_date)
    print("\n=== Regras painel ===")
    print("  lastActivityByChat: chatId -> {date ultimo pedido, chatName}")
    print("  Match: telefone planilha -> aliases BR -> chatId (messageDB/clientDB)")
    print("  Duplicata mesmo telefone: MAX(data ultimo pedido)")
    if dup:
        print("\n=== 3 exemplos duplicados ===")
        for dig, items in list(dup.items())[:3]:
            print(" ", dig, len(items), "linhas")
            for t in items:
                print("   ", t)

if __name__ == "__main__":
    main()

